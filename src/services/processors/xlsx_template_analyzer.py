"""
XLSX Template Analyzer - Extracts template structure for auto-fill mapping.

This analyzer reads an XLSX file and identifies:
- User input fields (editable cells that need data)
- Formula fields (auto-calculated, should not be overwritten)
- Label/header cells (static text for context)
- Data validation rules

Output: Template Schema JSON that maps semantic field names to cell locations.
"""

import json
import sys
import logging
import re
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('xlsx_template_analyzer')


class CellType:
    """Cell type constants."""
    FORMULA = "formula"
    INPUT = "input"
    LABEL = "label"
    EMPTY = "empty"
    HEADER = "header"


class XLSXTemplateAnalyzer:
    """
    Analyzes XLSX templates to extract structure for auto-filling.
    
    Features:
    - Identifies formula cells vs input cells
    - Detects labels and headers for context
    - Extracts data validation rules
    - Maps cell positions to semantic field names
    - Tracks merged cell regions
    """
    
    def __init__(self):
        """Initialize the analyzer."""
        self.workbook = None
        self.template_path = None
        
    def analyze(self, file_path: str) -> Dict[str, Any]:
        """
        Analyze an XLSX template file.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            Template schema dictionary with structure information
        """
        try:
            logger.info(f"Analyzing XLSX template: {file_path}")
            self.template_path = Path(file_path)
            
            # Load workbook with data_only=False to preserve formulas
            self.workbook = load_workbook(file_path, data_only=False)
            
            # Build schema
            schema = {
                "template_name": self.template_path.stem,
                "file_path": str(self.template_path),
                "analysis_date": datetime.now().isoformat(),
                "sheet_count": len(self.workbook.sheetnames),
                "sheets": []
            }
            
            # Analyze each sheet
            for sheet_name in self.workbook.sheetnames:
                logger.info(f"Analyzing sheet: {sheet_name}")
                sheet = self.workbook[sheet_name]
                sheet_schema = self._analyze_sheet(sheet)
                schema["sheets"].append(sheet_schema)
            
            # Generate summary statistics
            schema["summary"] = self._generate_summary(schema)
            
            logger.info(f"Analysis complete: {schema['summary']['total_input_fields']} input fields, "
                       f"{schema['summary']['total_formula_fields']} formula fields")
            
            return schema
            
        except Exception as e:
            logger.error(f"Error analyzing template: {str(e)}")
            return {
                "error": str(e),
                "template_name": Path(file_path).stem if file_path else "unknown"
            }
    
    def _analyze_sheet(self, sheet: Worksheet) -> Dict[str, Any]:
        """Analyze a single worksheet."""
        sheet_schema = {
            "name": sheet.title,
            "dimensions": sheet.dimensions,
            "max_row": sheet.max_row,
            "max_col": sheet.max_column,
            "input_fields": [],
            "formula_fields": [],
            "labels": [],
            "headers": [],
            "merged_regions": [],
            "data_validations": []
        }
        
        # Track merged cells
        for merged_range in sheet.merged_cells.ranges:
            sheet_schema["merged_regions"].append({
                "range": str(merged_range),
                "start_cell": merged_range.coord.split(":")[0]
            })
        
        # Track data validations
        if sheet.data_validations:
            for dv in sheet.data_validations.dataValidation:
                sheet_schema["data_validations"].append({
                    "type": dv.type,
                    "ranges": str(dv.sqref) if dv.sqref else None,
                    "formula1": str(dv.formula1) if dv.formula1 else None,
                    "formula2": str(dv.formula2) if dv.formula2 else None,
                    "allow_blank": dv.allowBlank
                })
        
        # Identify potential header rows (first few rows often contain headers)
        header_rows = self._detect_header_rows(sheet)
        
        # Analyze each cell
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_info = self._analyze_cell(cell, sheet, row_idx in header_rows)
                
                if cell_info:
                    # Categorize the cell
                    if cell_info["type"] == CellType.FORMULA:
                        sheet_schema["formula_fields"].append(cell_info)
                    elif cell_info["type"] == CellType.INPUT:
                        sheet_schema["input_fields"].append(cell_info)
                    elif cell_info["type"] == CellType.HEADER:
                        sheet_schema["headers"].append(cell_info)
                    elif cell_info["type"] == CellType.LABEL:
                        sheet_schema["labels"].append(cell_info)
        
        # Try to associate labels with input fields
        sheet_schema["input_fields"] = self._associate_labels(
            sheet_schema["input_fields"],
            sheet_schema["labels"],
            sheet_schema["headers"]
        )
        
        return sheet_schema
    
    def _analyze_cell(
        self, 
        cell: Cell, 
        sheet: Worksheet, 
        is_header_row: bool
    ) -> Optional[Dict[str, Any]]:
        """
        Analyze a single cell and determine its type and properties.
        
        Returns None for truly empty cells with no significance.
        """
        cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
        
        # Check if cell is part of a merged region
        is_merged = False
        merge_master = None
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                is_merged = True
                merge_master = merged_range.coord.split(":")[0]
                # Skip non-master cells in merged regions
                if cell.coordinate != merge_master:
                    return None
                break
        
        # Determine cell type
        cell_type = self._determine_cell_type(cell, is_header_row)
        
        if cell_type == CellType.EMPTY:
            return None
        
        # Build cell info
        cell_info = {
            "cell": cell_ref,
            "row": cell.row,
            "column": cell.column,
            "column_letter": get_column_letter(cell.column),
            "type": cell_type,
            "is_merged": is_merged
        }
        
        if merge_master:
            cell_info["merge_master"] = merge_master
        
        # Add type-specific information
        if cell_type == CellType.FORMULA:
            cell_info["formula"] = str(cell.value)
            cell_info["description"] = self._describe_formula(str(cell.value))
        elif cell_type in [CellType.LABEL, CellType.HEADER]:
            cell_info["text"] = str(cell.value) if cell.value else ""
        elif cell_type == CellType.INPUT:
            cell_info["current_value"] = cell.value
            cell_info["data_type"] = self._get_data_type(cell)
            cell_info["number_format"] = cell.number_format
            cell_info["label"] = None  # Will be filled by _associate_labels
            cell_info["json_path"] = None  # Will be filled during mapping
        
        # Add formatting info for context
        if cell.font:
            cell_info["is_bold"] = cell.font.bold
        if cell.fill and cell.fill.fgColor:
            cell_info["has_fill"] = cell.fill.fgColor.rgb != "00000000"
        
        return cell_info
    
    def _determine_cell_type(self, cell: Cell, is_header_row: bool) -> str:
        """Determine the type of a cell."""
        value = cell.value
        
        # Empty cell
        if value is None or (isinstance(value, str) and not value.strip()):
            return CellType.EMPTY
        
        # Formula cell
        if isinstance(value, str) and value.startswith("="):
            return CellType.FORMULA
        
        # Header row cells are headers
        if is_header_row and isinstance(value, str):
            return CellType.HEADER
        
        # Check if it's a label (text that describes adjacent cells)
        if isinstance(value, str):
            # Labels often end with ":" or are short descriptive text
            if value.endswith(":") or self._is_likely_label(value, cell):
                return CellType.LABEL
        
        # Check for input fields
        if self._is_likely_input(cell):
            return CellType.INPUT
        
        # Default to label for text, input for numbers
        if isinstance(value, str):
            return CellType.LABEL
        else:
            return CellType.INPUT
    
    def _is_likely_label(self, value: str, cell: Cell) -> bool:
        """
        Determine if a text value is likely a label.
        
        Labels are typically:
        - Short text (under 50 chars)
        - Bold formatted
        - In specific columns (A, B often)
        - End with colon
        - Common label words
        """
        if not value:
            return False
        
        # Common label patterns
        label_patterns = [
            r'^\s*[\w\s]+:\s*$',  # Ends with colon
            r'^(Total|Subtotal|Sum|Average|Count|Date|Name|Address|Property|Unit|Rent|Price|Rate|Fee|Cost|Amount|Number|#|No\.|Type|Status|Notes?|Description|Comments?)\s*:?\s*$',
        ]
        
        for pattern in label_patterns:
            if re.match(pattern, value, re.IGNORECASE):
                return True
        
        # Short text in column A or B is often a label
        if cell.column <= 2 and len(value) < 30:
            return True
        
        # Bold text is often a label
        if cell.font and cell.font.bold:
            return True
        
        return False
    
    def _is_likely_input(self, cell: Cell) -> bool:
        """
        Determine if a cell is likely an input field.
        
        Input fields are typically:
        - Numeric values without formulas
        - Date values
        - Cells with data validation
        - Cells in "data" columns (not A or B)
        """
        value = cell.value
        
        # Numbers are usually input
        if isinstance(value, (int, float)):
            return True
        
        # Dates are input
        if cell.is_date:
            return True
        
        # Cells with specific number formats
        if cell.number_format and cell.number_format != "General":
            return True
        
        return False
    
    def _detect_header_rows(self, sheet: Worksheet) -> set:
        """Detect which rows are likely header rows."""
        header_rows = set()
        
        # First row is often a header
        if sheet.max_row > 0:
            first_row_cells = [sheet.cell(row=1, column=c).value for c in range(1, min(10, sheet.max_column + 1))]
            if any(isinstance(v, str) and v for v in first_row_cells):
                header_rows.add(1)
        
        # Look for rows where most cells are bold text
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            bold_count = 0
            text_count = 0
            for col_idx in range(1, min(20, sheet.max_column + 1)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    text_count += 1
                    if cell.font and cell.font.bold:
                        bold_count += 1
            
            # If most text cells are bold, it's likely a header row
            if text_count > 0 and bold_count / text_count > 0.5:
                header_rows.add(row_idx)
        
        return header_rows
    
    def _associate_labels(
        self,
        input_fields: List[Dict],
        labels: List[Dict],
        headers: List[Dict]
    ) -> List[Dict]:
        """
        Associate labels with input fields based on proximity.
        
        Labels are typically directly to the left or above input fields.
        """
        for field in input_fields:
            row = field["row"]
            col = field["column"]
            
            # Look for label to the left
            left_label = self._find_label_at(labels, row, col - 1)
            if left_label:
                field["label"] = left_label["text"]
                field["label_cell"] = left_label["cell"]
                continue
            
            # Look for label above
            above_label = self._find_label_at(labels, row - 1, col)
            if above_label:
                field["label"] = above_label["text"]
                field["label_cell"] = above_label["cell"]
                continue
            
            # Look for header in same column
            col_header = next(
                (h for h in headers if h["column"] == col),
                None
            )
            if col_header:
                field["label"] = col_header["text"]
                field["label_cell"] = col_header["cell"]
        
        return input_fields
    
    def _find_label_at(
        self,
        labels: List[Dict],
        row: int,
        col: int
    ) -> Optional[Dict]:
        """Find a label at a specific position."""
        for label in labels:
            if label["row"] == row and label["column"] == col:
                return label
        return None
    
    def _get_data_type(self, cell: Cell) -> str:
        """Determine the expected data type for an input cell."""
        value = cell.value
        number_format = cell.number_format or "General"
        
        # Check number format first
        if "$" in number_format or "Currency" in number_format:
            return "currency"
        if "%" in number_format:
            return "percentage"
        if any(fmt in number_format.lower() for fmt in ["d", "m", "y"]):
            return "date"
        
        # Check value type
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, int):
            return "integer"
        if isinstance(value, float):
            return "number"
        if cell.is_date:
            return "date"
        
        return "text"
    
    def _describe_formula(self, formula: str) -> str:
        """Generate a human-readable description of a formula."""
        # Common formula patterns
        if formula.upper().startswith("=SUM("):
            return "Sum calculation"
        if formula.upper().startswith("=AVERAGE("):
            return "Average calculation"
        if formula.upper().startswith("=IF("):
            return "Conditional calculation"
        if formula.upper().startswith("=VLOOKUP(") or formula.upper().startswith("=XLOOKUP("):
            return "Lookup reference"
        if "*" in formula:
            return "Multiplication calculation"
        if "/" in formula:
            return "Division calculation"
        if "+" in formula:
            return "Addition calculation"
        if "-" in formula:
            return "Subtraction calculation"
        
        return "Formula calculation"
    
    def _generate_summary(self, schema: Dict) -> Dict[str, Any]:
        """Generate summary statistics for the schema."""
        total_input = 0
        total_formula = 0
        total_labels = 0
        sheets_summary = []
        
        for sheet in schema["sheets"]:
            input_count = len(sheet["input_fields"])
            formula_count = len(sheet["formula_fields"])
            label_count = len(sheet["labels"])
            
            total_input += input_count
            total_formula += formula_count
            total_labels += label_count
            
            sheets_summary.append({
                "name": sheet["name"],
                "input_fields": input_count,
                "formula_fields": formula_count,
                "labels": label_count
            })
        
        return {
            "total_input_fields": total_input,
            "total_formula_fields": total_formula,
            "total_labels": total_labels,
            "sheets": sheets_summary
        }


def main():
    """CLI entry point for analyzing XLSX templates."""
    if len(sys.argv) < 2:
        print(json.dumps({
            "error": "Usage: python xlsx_template_analyzer.py <xlsx_path> [output_path]"
        }))
        sys.exit(1)
    
    xlsx_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    # Check file exists
    if not Path(xlsx_path).exists():
        print(json.dumps({
            "error": f"File not found: {xlsx_path}"
        }))
        sys.exit(1)
    
    # Analyze the template
    analyzer = XLSXTemplateAnalyzer()
    schema = analyzer.analyze(xlsx_path)
    
    # Output result
    if output_path:
        with open(output_path, 'w') as f:
            json.dump(schema, f, indent=2, default=str)
        print(json.dumps({"status": "success", "output_path": output_path}))
    else:
        print(json.dumps(schema, indent=2, default=str))


if __name__ == "__main__":
    main()

