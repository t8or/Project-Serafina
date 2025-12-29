"""
XLSX Template Filler - Populates XLSX templates with data from JSON extracts.

This filler reads a template schema and field mappings, then populates
an XLSX template with data extracted from PDF documents via Docling.

Features:
- Only fills user-input cells (preserves formulas)
- Applies data transformations (percentage conversion, etc.)
- Validates data types before insertion
- Tracks which fields were filled vs. need manual input
- Saves timestamped output files
"""

import json
import sys
import logging
import re
import shutil
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('xlsx_template_filler')


class XLSXTemplateFiller:
    """
    Fills XLSX templates with data from JSON extracts.
    
    Features:
    - Uses field mappings to locate target cells
    - Applies transformations (percentage, currency, etc.)
    - Preserves formulas and formatting
    - Generates fill report showing what was populated
    """
    
    def __init__(self, mappings_path: str = None):
        """
        Initialize the filler with optional mappings file.
        
        Args:
            mappings_path: Path to field_mappings.json
        """
        self.mappings = None
        if mappings_path:
            self.load_mappings(mappings_path)
    
    def load_mappings(self, mappings_path: str) -> None:
        """Load field mappings from JSON file."""
        with open(mappings_path, 'r') as f:
            self.mappings = json.load(f)
        logger.info(f"Loaded field mappings: {self.mappings.get('template_name', 'unknown')}")
    
    def fill(
        self,
        template_path: str,
        json_data: Dict[str, Any],
        output_path: str = None,
        mappings: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """
        Fill an XLSX template with data from a JSON extract.
        
        Args:
            template_path: Path to the XLSX template file
            json_data: Extracted JSON data from Docling
            output_path: Optional output path (auto-generated if not provided)
            mappings: Optional field mappings (uses loaded mappings if not provided)
            
        Returns:
            Fill report with statistics and any errors
        """
        try:
            # Use provided mappings or loaded mappings
            field_mappings = mappings or self.mappings
            if not field_mappings:
                return {
                    "status": "error",
                    "error": "No field mappings provided or loaded"
                }
            
            logger.info(f"Filling template: {template_path}")
            
            # Copy template to output location
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                template_name = Path(template_path).stem
                output_path = str(Path(template_path).parent / f"{template_name}_filled_{timestamp}.xlsx")
            
            shutil.copy2(template_path, output_path)
            
            # Load the copied workbook
            workbook = load_workbook(output_path)
            
            # Extract the root data from JSON
            json_root = field_mappings.get("json_root", "structured_data[0]")
            data = self._navigate_json_path(json_data, json_root)
            
            if data is None:
                return {
                    "status": "error",
                    "error": f"Could not find data at json_root: {json_root}"
                }
            
            # Track fill results
            fill_report = {
                "status": "success",
                "output_path": output_path,
                "filled_fields": [],
                "skipped_fields": [],
                "external_fields": [],
                "errors": [],
                "timestamp": datetime.now().isoformat()
            }
            
            # Process each sheet's mappings
            sheets_config = field_mappings.get("sheets", {})
            
            for sheet_name, sheet_config in sheets_config.items():
                if sheet_name not in workbook.sheetnames:
                    logger.warning(f"Sheet not found in template: {sheet_name}")
                    fill_report["errors"].append({
                        "sheet": sheet_name,
                        "error": "Sheet not found in template"
                    })
                    continue
                
                sheet = workbook[sheet_name]
                
                # Process single-cell mappings
                if "mappings" in sheet_config:
                    for mapping in sheet_config["mappings"]:
                        result = self._fill_cell(sheet, mapping, data, json_data)
                        self._add_to_report(fill_report, sheet_name, mapping, result)
                
                # Process array-based mappings (like unit mix)
                if "array_source" in sheet_config:
                    self._fill_array_data(sheet, sheet_config, data, fill_report)
            
            # Save the filled workbook
            workbook.save(output_path)
            logger.info(f"Saved filled template to: {output_path}")
            
            # Generate summary
            fill_report["summary"] = {
                "total_filled": len(fill_report["filled_fields"]),
                "total_skipped": len(fill_report["skipped_fields"]),
                "total_external": len(fill_report["external_fields"]),
                "total_errors": len(fill_report["errors"])
            }
            
            return fill_report
            
        except Exception as e:
            logger.error(f"Error filling template: {str(e)}")
            return {
                "status": "error",
                "error": str(e)
            }
    
    def _fill_cell(
        self,
        sheet,
        mapping: Dict[str, Any],
        data: Dict[str, Any],
        full_json: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """
        Fill a single cell based on mapping configuration.
        
        Returns result dict with status and details.
        """
        cell_ref = mapping.get("cell")
        json_path = mapping.get("json_path")
        source = mapping.get("source", "pdf_extract")
        json_root_override = mapping.get("json_root_override")
        
        # Handle external fields (no json_path)
        if source == "external" or json_path is None:
            return {
                "status": "external",
                "cell": cell_ref,
                "label": mapping.get("label"),
                "notes": mapping.get("notes")
            }
        
        # Get value from JSON data
        # Use full_json if json_root_override is set (empty string means use root)
        lookup_data = full_json if (json_root_override is not None and full_json) else data
        value = self._navigate_json_path(lookup_data, json_path)
        
        # Try fallback path if primary path didn't find a value
        fallback_path = mapping.get("fallback_path")
        fallback_transform = mapping.get("fallback_transform")
        if value is None and fallback_path and full_json:
            value = self._navigate_json_path(full_json, fallback_path)
            # Apply fallback transform if specified
            if value is not None and fallback_transform:
                value = self._apply_transform(value, fallback_transform)
        
        if value is None:
            return {
                "status": "skipped",
                "cell": cell_ref,
                "label": mapping.get("label"),
                "reason": f"No value found at path: {json_path}"
            }
        
        # Apply transformation if specified
        transform = mapping.get("transform")
        if transform:
            value = self._apply_transform(value, transform)
        
        # Set the cell value
        try:
            cell = sheet[cell_ref]
            
            # Check if cell has a formula (don't overwrite)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                return {
                    "status": "skipped",
                    "cell": cell_ref,
                    "label": mapping.get("label"),
                    "reason": "Cell contains formula, not overwriting"
                }
            
            cell.value = value
            
            return {
                "status": "filled",
                "cell": cell_ref,
                "label": mapping.get("label"),
                "value": value,
                "json_path": json_path
            }
            
        except Exception as e:
            return {
                "status": "error",
                "cell": cell_ref,
                "label": mapping.get("label"),
                "error": str(e)
            }
    
    def _fill_array_data(
        self,
        sheet,
        sheet_config: Dict[str, Any],
        data: Dict[str, Any],
        fill_report: Dict[str, Any]
    ) -> None:
        """Fill array-based data like unit mix rows."""
        array_source = sheet_config.get("array_source")
        row_start = sheet_config.get("row_start", 1)
        column_mappings = sheet_config.get("mappings", [])
        max_rows = sheet_config.get("max_rows", 25)  # Max rows to clear in template
        
        # Clear the entire array range first (to remove stale template data)
        clear_columns = sheet_config.get("clear_columns", [col.get("column") for col in column_mappings if col.get("column")])
        delete_empty_rows = sheet_config.get("delete_empty_rows", True)
        
        # Find the "Total" row - this marks the end of unit mix data area
        # We should only clear/delete rows BEFORE the Total row
        total_row = None
        last_unit_row = row_start
        for check_row in range(row_start, row_start + 50):
            cell_val = sheet[f"B{check_row}"].value
            if cell_val is not None:
                cell_str = str(cell_val).strip().lower()
                if cell_str == "total" or cell_str == "totals":
                    total_row = check_row
                    break
                elif cell_str:  # Non-empty, non-total = unit row
                    last_unit_row = check_row
        
        # If no Total row found, use max_rows from config
        if total_row:
            # Only clear rows from row_start to total_row - 1 (don't touch Total row or below)
            rows_in_table = total_row - row_start
        else:
            rows_in_table = max_rows
        
        max_rows = rows_in_table  # Only process rows up to (but not including) Total row
        
        for column in clear_columns:
            if column:
                for row_offset in range(max_rows):
                    clear_row = row_start + row_offset
                    cell_ref = f"{column}{clear_row}"
                    try:
                        cell = sheet[cell_ref]
                        # Only clear if not a formula
                        if not (cell.value and isinstance(cell.value, str) and cell.value.startswith("=")):
                            cell.value = None
                    except:
                        pass
        
        # Track how many actual data rows we'll fill (for later row deletion)
        rows_to_fill = 0
        
        # Get the array data
        array_data = self._navigate_json_path(data, array_source)
        
        if not array_data or not isinstance(array_data, list):
            logger.warning(f"No array data found at: {array_source}")
            return
        
        sheet_name = sheet.title
        
        # Process each row of data
        for row_idx, row_data in enumerate(array_data):
            current_row = row_start + row_idx
            
            # Skip summary/total rows (Totals, All X Beds, etc.)
            bed_value = row_data.get("bed", "")
            if isinstance(bed_value, str):
                bed_lower = bed_value.lower()
                if "total" in bed_lower or "all" in bed_lower:
                    continue
            
            rows_to_fill += 1
            
            for col_mapping in column_mappings:
                column = col_mapping.get("column")
                json_field = col_mapping.get("json_field")
                transform = col_mapping.get("transform")
                
                if not column:
                    continue
                
                # Handle special transforms
                if transform == "bed_bath_label":
                    # Create unit type label like "2B/1Ba" from bed and bath values
                    bed = row_data.get("bed")
                    bath = row_data.get("bath")
                    if isinstance(bed, (int, float)) and isinstance(bath, (int, float)):
                        value = f"{int(bed)}B/{int(bath)}Ba"
                    else:
                        continue
                elif transform == "calc_occupied":
                    # Calculate occupied units = total units - available units
                    total_units = self._navigate_json_path(row_data, "unitMix.units")
                    available_units = self._navigate_json_path(row_data, "availability.units")
                    if total_units is not None and available_units is not None:
                        value = int(total_units) - int(available_units)
                    else:
                        continue
                elif not json_field:
                    continue
                else:
                    # Get value from row data
                    value = self._navigate_json_path(row_data, json_field)
                
                if value is None:
                    continue
                
                cell_ref = f"{column}{current_row}"
                
                try:
                    cell = sheet[cell_ref]
                    
                    # Don't overwrite formulas
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        continue
                    
                    cell.value = value
                    
                    fill_report["filled_fields"].append({
                        "sheet": sheet_name,
                        "cell": cell_ref,
                        "label": col_mapping.get("label"),
                        "value": value
                    })
                    
                except Exception as e:
                    fill_report["errors"].append({
                        "sheet": sheet_name,
                        "cell": cell_ref,
                        "error": str(e)
                    })
        
        # Delete empty rows after filling (delete from bottom to top to avoid index shifting)
        if delete_empty_rows and rows_to_fill < max_rows:
            rows_to_delete = max_rows - rows_to_fill
            first_empty_row = row_start + rows_to_fill
            
            # Delete rows from the first empty row
            sheet.delete_rows(first_empty_row, rows_to_delete)
    
    def _navigate_json_path(self, data: Any, path: str) -> Any:
        """
        Navigate a JSON path like 'property.no_of_units' or 'unitBreakdown[1].rows'.
        
        Supports:
        - Direct key lookup (for keys containing dots like 'unitMix.units')
        - Dot notation: property.no_of_units
        - Array indexing: unitBreakdown[1]
        - Mixed: structured_data[0].property.name
        """
        if data is None or not path:
            return None
        
        # FIRST: Try direct key lookup (handles keys with dots in name like "unitMix.units")
        if isinstance(data, dict) and path in data:
            return data[path]
        
        current = data
        
        # Split path by dots, but handle array notation
        parts = re.split(r'\.(?![^\[]*\])', path)
        
        for part in parts:
            if current is None:
                return None
            
            # Check for array indexing
            array_match = re.match(r'(\w+)\[(\d+)\]', part)
            
            if array_match:
                key = array_match.group(1)
                index = int(array_match.group(2))
                
                if isinstance(current, dict) and key in current:
                    current = current[key]
                    if isinstance(current, list) and index < len(current):
                        current = current[index]
                    else:
                        return None
                elif isinstance(current, list) and key.isdigit():
                    idx = int(key)
                    if idx < len(current):
                        current = current[idx]
                    else:
                        return None
                else:
                    return None
            else:
                # Regular key access
                if isinstance(current, dict):
                    current = current.get(part)
                else:
                    return None
        
        return current
    
    def _apply_transform(self, value: Any, transform: str) -> Any:
        """Apply a transformation to a value."""
        if transform == "divide_by_100":
            # Convert percentage from whole number to decimal
            try:
                return float(value) / 100
            except (ValueError, TypeError):
                return value
        
        elif transform == "to_number":
            # Convert string to number, removing formatting
            if isinstance(value, str):
                try:
                    cleaned = re.sub(r'[^0-9.-]', '', value)
                    return float(cleaned)
                except ValueError:
                    return value
            return value
        
        elif transform == "years_since_purchase":
            # Extract year from "Purchased Mar 2019" and calculate years
            if isinstance(value, str):
                year_match = re.search(r'\b(19|20)\d{2}\b', value)
                if year_match:
                    purchase_year = int(year_match.group())
                    current_year = datetime.now().year
                    return current_year - purchase_year
            return value
        
        elif transform == "extract_property_name":
            # Extract property name from "Urban 148 148 Unit Apartment Building"
            # Take text before "Unit" or first number sequence
            if isinstance(value, str):
                # Try to find pattern like "Name ### Unit"
                match = re.match(r'^(.+?)\s+\d+\s+Unit', value)
                if match:
                    return match.group(1).strip()
                # Fallback: take first part before numbers
                match = re.match(r'^([A-Za-z\s]+)', value)
                if match:
                    return match.group(1).strip()
            return value
        
        elif transform == "extract_city":
            # Extract city from "Phoenix, Arizona - North Phoenix Neighborhood"
            if isinstance(value, str):
                # Take text before first comma
                parts = value.split(',')
                if parts:
                    return parts[0].strip()
            return value
        
        elif transform == "extract_state_abbrev":
            # Extract state abbreviation from "Phoenix, Arizona - North Phoenix Neighborhood"
            state_abbrevs = {
                'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR',
                'california': 'CA', 'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE',
                'florida': 'FL', 'georgia': 'GA', 'hawaii': 'HI', 'idaho': 'ID',
                'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA', 'kansas': 'KS',
                'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
                'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN', 'mississippi': 'MS',
                'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE', 'nevada': 'NV',
                'new hampshire': 'NH', 'new jersey': 'NJ', 'new mexico': 'NM', 'new york': 'NY',
                'north carolina': 'NC', 'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK',
                'oregon': 'OR', 'pennsylvania': 'PA', 'rhode island': 'RI', 'south carolina': 'SC',
                'south dakota': 'SD', 'tennessee': 'TN', 'texas': 'TX', 'utah': 'UT',
                'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA', 'west virginia': 'WV',
                'wisconsin': 'WI', 'wyoming': 'WY'
            }
            if isinstance(value, str):
                # Look for state name after comma, before dash or end
                match = re.search(r',\s*([A-Za-z\s]+?)(?:\s*-|$)', value)
                if match:
                    state_name = match.group(1).strip().lower()
                    return state_abbrevs.get(state_name, state_name.upper()[:2])
            return value
        
        return value
    
    def _add_to_report(
        self,
        report: Dict[str, Any],
        sheet_name: str,
        mapping: Dict[str, Any],
        result: Dict[str, Any]
    ) -> None:
        """Add a fill result to the report."""
        result["sheet"] = sheet_name
        
        status = result.get("status")
        
        if status == "filled":
            report["filled_fields"].append(result)
        elif status == "skipped":
            report["skipped_fields"].append(result)
        elif status == "external":
            report["external_fields"].append(result)
        elif status == "error":
            report["errors"].append(result)


def main():
    """CLI entry point for filling XLSX templates."""
    if len(sys.argv) < 4:
        print(json.dumps({
            "error": "Usage: python xlsx_template_filler.py <template_path> <json_path> <mappings_path> [output_path]"
        }))
        sys.exit(1)
    
    template_path = sys.argv[1]
    json_path = sys.argv[2]
    mappings_path = sys.argv[3]
    output_path = sys.argv[4] if len(sys.argv) > 4 else None
    
    # Validate files exist
    for path, name in [(template_path, "Template"), (json_path, "JSON"), (mappings_path, "Mappings")]:
        if not Path(path).exists():
            print(json.dumps({
                "error": f"{name} file not found: {path}"
            }))
            sys.exit(1)
    
    # Load JSON data
    with open(json_path, 'r') as f:
        json_data = json.load(f)
    
    # Create filler and fill template
    filler = XLSXTemplateFiller(mappings_path)
    result = filler.fill(template_path, json_data, output_path)
    
    # Output result
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()

